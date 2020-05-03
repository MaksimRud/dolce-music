from django.shortcuts import render
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.template.response import TemplateResponse
from django.http import HttpResponse, request, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views import generic
<<<<<<< HEAD
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.template.defaulttags import register

import openpyxl
import io
import xlsxwriter
from django.utils.translation import ugettext
from .fusioncharts import FusionCharts
import re

from .forms import *
from .models import *
from django.db.models import Count
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .decorators import admin_required
=======
from .forms import *
from .models import *
from .fusioncharts import FusionCharts
#from .forms import PeriodEditFrom
# Create your views here.
>>>>>>> 3c85a009d41e0aa88e8966791cf5bf3e6148b68e


@login_required
def index(request):
    return TemplateResponse(request, 'search.html', {})


def home_view(request):
    period_values = Period.objects.order_by('time_of_period').values('name')
    context = {'period_names': period_values}

    return render(request, 'music_search/index.html', context)

def validate_period(value):
    characters = value.split('-')
    check = True

    if len(characters) != 2:
        return False

    first_date = characters[0]
    second_date = characters[1]

    if not first_date.isdigit() or not second_date.isdigit():
        return False
    else:
        first_date = int(first_date)
        second_date = int(second_date)

    if first_date >= second_date:
        check = False

    if second_date - first_date > 500:
        check = False

    return check

def validate_date(value):
    check = False
    if re.match(r"([12]\d{3}-(0[1-9]|1[0-2])-(0[1-9]|[12]\d|3[01]))", value):
        check = True
    return check

class PeriodView(LoginRequiredMixin, generic.ListView):
    model = Period
    template_name = 'music_search/periods.html'
    login_url = '/login/'

    def get_context_data(self, **kwargs):
        context = super(PeriodView, self).get_context_data(**kwargs)

        period_values = Period.objects.order_by(
            'time_of_period').values('time_of_period', 'name', 'id')
        context['period_list'] = Period.objects.order_by('time_of_period')
        context['header'] = list(period_values[0].keys())[:-1]
        context['content_period'] = list(period_values)
        return context

@admin_required
def post_per_file(request):
    periods = Period.objects.all()
    period_values = Period.objects.order_by(
        'time_of_period').values('time_of_period', 'name', 'descr', 'id')
    
    template_name = 'music_search/periods.html'

    if request.method == 'POST' and request.FILES['excel-file']:
        excel_file = request.FILES['excel-file']
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        uploaded_file_url = fs.url(filename)

<<<<<<< HEAD
        wb = openpyxl.load_workbook(excel_file)

        worksheet = wb[wb.sheetnames[0]]

        period_name = [period['name'] for period in period_values]
        period_year = [period['time_of_period'] for period in period_values]
        period_descr = [period['descr'] for period in period_values]

        
        i = 0
        for data in worksheet.iter_rows(min_row=1, max_col=4, values_only=True):
    
            if i == 0:
                i += 1
                continue
            i += 1

            name = data[1]

            if name == None:
                break

            elif name not in period_name:

                if (data[2] in period_year or data[3] in period_descr):
                    continue
                elif not validate_period(data[2]):
                    continue
                else:

                    new_model = {
                        'name': data[1], 
                        'time_of_period': data[2], 
                        'descr': data[3]
                    }

                    p = Period(**new_model)
                    p.save()
            else:
                for period in period_values:
                    if name != period['name']:
                        continue

                    else:
                        if (data[2] in period_values or data[3] in period_values):
                            continue
                        if validate_period(data[2]):
    
                            time_of_period = data[2]
                            descr = data[3]

                            Period.objects.filter(pk=period['id']).update(
                                time_of_period=time_of_period, descr=descr)

    period_values = Period.objects.order_by(
        'time_of_period').values('time_of_period', 'name', 'id')
    header = list(period_values[0].keys())[:-1]
    content_period = list(period_values)
    period_list = Period.objects.order_by('time_of_period')

                
    return render(request, 'music_search/periods.html', {
        'uploaded_file_url': uploaded_file_url,
        'header': header,
        'content_period': content_period,
        'period_list': period_list
    })


def get_per_file(request):
    periods = Period.objects.all()
    template_name = 'music_search/periods.html'

    output = io.BytesIO()

    workbook = xlsxwriter.Workbook(output)
    worksheet_s = workbook.add_worksheet("Summary")
    title = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
        'valign': 'vcenter'
    })
    header = workbook.add_format({
        'bg_color': '#F7F7F7',
        'color': 'black',
        'align': 'center',
        'valign': 'top',
        'border': 1
    })
    cell = workbook.add_format({
        'align': 'left',
        'valign': 'top',
        'text_wrap': True,
        'border': 1
    })
    cell_center = workbook.add_format({
        'align': 'center',
        'valign': 'top',
        'border': 1
    })

    worksheet_s.write(0, 0, ugettext("No"), header)
    worksheet_s.write(0, 1, ugettext("name"), header)
    worksheet_s.write(0, 2, ugettext("time"), header)
    worksheet_s.write(0, 3, ugettext("descr"), header)

    description_col_width = 10

    for idx, data in enumerate(periods):
        row = 1 + idx
        worksheet_s.write_number(row, 0, idx + 1, cell_center)
        worksheet_s.write_string(row, 1, data.name, cell)
        worksheet_s.write_string(row, 2, data.time_of_period, cell_center)
        worksheet_s.write_string(row, 3, data.descr, cell)
        if len(data.descr) > description_col_width:
            description_col_width = len(data.descr)

    workbook.close()
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
    return response


class CompousersView(LoginRequiredMixin, generic.ListView):
    model = Compouser
    template_name = 'music_search/compousers.html'
    login_url = '/login/'
=======
class CompousersView(generic.ListView):
    model = Compousers
    template_name = 'music_search/compousers.html'
>>>>>>> 3c85a009d41e0aa88e8966791cf5bf3e6148b68e

    def chart(self):
        dataSource = {}
        dataSource['chart'] = {
            "caption": "Compousers in Periods chart",
            "numberPrefix": "$",
            "bgColor": "#ffffff",
            "startingAngle": "310",
            "centerLabelBold": "1",
            "showTooltip": "0",
            "decimals": "0",
            "theme": "fusion"
        }
<<<<<<< HEAD

        dataSource['data'] = []

        for key in Period.objects.annotate(num_compousers=Count('compouser')):
            data = {}
            data['label'] = key.name
            data['value'] = key.num_compousers
=======
        dataSource['data'] = []

        Renesans_comp = 0
        Baroque_comp = 0
        Classical_comp = 0
        Romatic_comp = 0
        Postmod = 0

        for key in Compousers.objects.all():
            if key.period.name == 'Baroque':
                Baroque_comp += 1
            elif key.period.name == 'Renaissance':
                Renesans_comp += 1
            elif key.period.name == 'Classical':
                Classical_comp += 1
            elif key.period.name == 'Romantic':
                Romatic_comp += 1

        periods = [Classical_comp, Baroque_comp,
                   Romatic_comp, Renesans_comp, Postmod]

        index = 0
        for key in Period.objects.all():
            data = {}
            data['label'] = key.name
            data['value'] = periods[index]
            index += 1
>>>>>>> 3c85a009d41e0aa88e8966791cf5bf3e6148b68e
            dataSource['data'].append(data)

        pie2D = FusionCharts(type="doughnut2d", id="ex1", width="550", height="450",
                             renderAt="chart-container", dataFormat="json", dataSource=dataSource)

        return pie2D
<<<<<<< HEAD
=======

    def get_context_data(self, **kwargs):
        context = super(CompousersView, self).get_context_data(**kwargs)
        compouser_values = Compousers.objects.order_by(
            'birth_date').values('name', 'birth_date', 'death_date', 'id')
        index = 0
        for comp in compouser_values:
            c = Compousers.objects.get(pk=comp['id'])
            compouser_values[index]['period'] = c.period.name
            del compouser_values[index]['id']
            compouser_values[index]['id'] = c.id
            index += 1
        context['compousers_list'] = Compousers.objects.order_by('birth_date')
        context['header'] = list(compouser_values[0].keys())[:-1]
        context['content_compousers'] = list(compouser_values)
        context['chart_compousers'] = self.chart().render()
        return context


class DetailView(generic.DetailView):
    model = Period
    template_name = 'music_search/period_detail.html'
>>>>>>> 3c85a009d41e0aa88e8966791cf5bf3e6148b68e

    def get_context_data(self, **kwargs):
        context = super(CompousersView, self).get_context_data(**kwargs)
        compouser_values = Compouser.objects.order_by(
            'birth_date').values('name', 'birth_date', 'death_date', 'id')
        index = 0

        for comp in compouser_values:
            c = Compouser.objects.get(pk=comp['id'])
            compouser_values[index]['period'] = c.period.name
            del compouser_values[index]['id']
            compouser_values[index]['id'] = c.id
            index += 1

        context['compousers_list'] = Compouser.objects.order_by('birth_date')
        context['header'] = list(compouser_values[0].keys())[:-1]
        context['content_compousers'] = list(compouser_values)
        context['chart_compousers'] = self.chart().render()

        return context

def get_comp_file(request):
    comp = Compouser.objects.all()

    template_name = 'music_search/compousers.html'

    output = io.BytesIO()

    workbook = xlsxwriter.Workbook(output)
    worksheet_s = workbook.add_worksheet("Summary")
    title = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
        'valign': 'vcenter'
    })
    header = workbook.add_format({
        'bg_color': '#F7F7F7',
        'color': 'black',
        'align': 'center',
        'valign': 'top',
        'border': 1
    })
    cell = workbook.add_format({
        'align': 'left',
        'valign': 'top',
        'text_wrap': True,
        'border': 1
    })
    cell_center = workbook.add_format({
        'align': 'center',
        'valign': 'top',
        'border': 1
    })

    worksheet_s.write(0, 0, ugettext("No"), header)
    worksheet_s.write(0, 1, ugettext("name"), header)
    worksheet_s.write(0, 2, ugettext("birth_date"), header)
    worksheet_s.write(0, 3, ugettext("death_date"), header)
    worksheet_s.write(0, 4, ugettext("period"), header)

    description_col_width = 10

    for idx, data in enumerate(comp):
        row = 1 + idx
        comp = Compouser.objects.get(pk=data.id)
        worksheet_s.write_number(row, 0, idx + 1, cell_center)
        worksheet_s.write_string(row, 1, data.name, cell)
        worksheet_s.write_string(row, 2, str(data.birth_date), cell_center)
        worksheet_s.write_string(row, 3, str(data.death_date), cell)
        worksheet_s.write_string(row, 4, comp.period.name, cell)

    workbook.close()
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
    return response

@admin_required
def post_comp_file(request):
    comp = Compouser.objects.all()
    comp_values = Compouser.objects.order_by(
        'birth_date').values('name', 'birth_date', 'death_date', 'id')
    period_values = Period.objects.order_by(
            'time_of_period').values('time_of_period', 'name', 'id')
    template_name = 'music_search/compousers.html'

    if request.method == 'POST' and request.FILES['excel-file']:
        excel_file = request.FILES['excel-file']
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        uploaded_file_url = fs.url(filename)
        wb = openpyxl.load_workbook(excel_file)
        
        worksheet = wb[wb.sheetnames[0]]
        

        compouser = [comp['name'] for comp in comp_values]
        
        
        i = 0

        for data in worksheet.iter_rows(min_row=1, max_col=5, values_only=True):

            if i == 0:
                i += 1
                continue

            i += 1

            name = data[1]
            period_name = [period['name'] for period in period_values]

            if name == None:
                break
            elif name not in compouser:
            
                if data[4] not in period_name  or not validate_date(str(data[2])) or not validate_date(str(data[3])):
                    continue

                p = Period.objects.get(name = data[4])

                new_model = {
                    'name': data[1], 
                    'birth_date': data[2], 
                    'death_date': data[3], 
                    'period': p 
                }

                c = Compouser(**new_model)
                c.save()

            else:
        
                for c in comp_values:

                    if name != c['name']:
                        continue

                    else:
                        birth_date = data[2]
                        death_date = data[3]

                        if data[4] not in period_name or not validate_date(str(birth_date)) or not validate_date(str(death_date)):
                            continue

                        period = data[4]
                        p = Period.objects.get(name = period)

                        Compouser.objects.filter(pk=c['id']).update(
                            birth_date=birth_date, 
                            death_date=death_date, 
                            period=p
                        )

    compouser_values = Compouser.objects.order_by(
        'birth_date').values('name', 'birth_date', 'death_date', 'id')

    index = 0
    for comp in compouser_values:
        c = Compouser.objects.get(pk=comp['id'])
        compouser_values[index]['period'] = c.period.name
        del compouser_values[index]['id']
        compouser_values[index]['id'] = c.id
        index += 1

    header = list(compouser_values[0].keys())[:-1]
    content_compouser = list(compouser_values)
    print(content_compouser)
    compouser_list = Compouser.objects.order_by('birth_date')
    
    return render(request, 'music_search/compousers.html', {
        'uploaded_file_url': uploaded_file_url,
        'header': header,
        'content_compousers': content_compouser,
        'compousers_list': compouser_list
    })

class PieceOfMusicView(LoginRequiredMixin, generic.ListView):
    model = PieceOfMusic
    template_name = 'music_search/music.html'
    login_url = '/login/'

    def chart(self):
        dataSource = {}
        dataSource['chart'] = {
            "caption": "Compousers in Periods chart",
            "numberPrefix": "$",
            "bgColor": "#ffffff",
            "startingAngle": "310",
            "centerLabelBold": "1",
            "showTooltip": "0",
            "decimals": "0",
            "theme": "fusion"
        }
        dataSource['data'] = []

        for key in Compouser.objects.annotate(num_music=Count('pieceofmusic')):
            data = {}
            data['label'] = key.name
            data['value'] = key.num_music
            dataSource['data'].append(data)
              
        pie2D = FusionCharts(type="doughnut2d", id="ex1", width="550", height="450",
                             renderAt="chart-container", dataFormat="json", dataSource=dataSource)

        return pie2D

    def get_context_data(self, **kwargs):
        context = super(PieceOfMusicView, self).get_context_data(**kwargs)
        music_values = PieceOfMusic.objects.order_by(
            'year_written').values('name', 'year_written', 'id')
        index = 0
        
        for mus in music_values:
            m = PieceOfMusic.objects.get(pk=mus['id'])
            music_values[index]['compouser'] = m.compousers.name
            music_values[index]['type_of_piece'] = m.type_of_piece.name
            music_values[index]['year_written'] = str(m.year_written)[:4]
            del music_values[index]['id']
            music_values[index]['id'] = m.id
            index += 1

        context['music_list'] = PieceOfMusic.objects.order_by('year_written')
        context['header'] = list(music_values[0].keys())[:-1]
        context['content_music'] = list(music_values)
        context['chart_music'] = self.chart().render()

        return context

def get_music_file(request):
    music = PieceOfMusic.objects.all()

    template_name = 'music_search/music.html'

    output = io.BytesIO()

    workbook = xlsxwriter.Workbook(output)
    worksheet_s = workbook.add_worksheet("Summary")
    title = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
        'valign': 'vcenter'
    })
    header = workbook.add_format({
        'bg_color': '#F7F7F7',
        'color': 'black',
        'align': 'center',
        'valign': 'top',
        'border': 1
    })
    cell = workbook.add_format({
        'align': 'left',
        'valign': 'top',
        'text_wrap': True,
        'border': 1
    })
    cell_center = workbook.add_format({
        'align': 'center',
        'valign': 'top',
        'border': 1
    })

    worksheet_s.write(0, 0, ugettext("No"), header)
    worksheet_s.write(0, 1, ugettext("name"), header)
    worksheet_s.write(0, 2, ugettext("year_written"), header)
    worksheet_s.write(0, 3, ugettext("compouser"), header)
    worksheet_s.write(0, 4, ugettext("type"), header)


    description_col_width = 10

    for idx, data in enumerate(music):
        row = 1 + idx
        mus = PieceOfMusic.objects.get(pk=data.id)
        worksheet_s.write_number(row, 0, idx + 1, cell_center)
        worksheet_s.write_string(row, 1, data.name, cell)
        worksheet_s.write_string(row, 2, str(data.year_written), cell_center)
        worksheet_s.write_string(row, 3, mus.compousers.name, cell)
        worksheet_s.write_string(row, 4, mus.type_of_piece.name, cell)


    workbook.close()
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
    return response

@admin_required
def post_music_file(request):

    music = PieceOfMusic.objects.all()
    music_values = PieceOfMusic.objects.order_by(
            'year_written').values('name', 'year_written', 'compousers', 'id')
    comp_values = Compouser.objects.values('name')
    type_ = TypeOfPiece.objects.values('name')

    template_name = 'music_search/music.html'
    
    if request.method == 'POST' and request.FILES['excel-file']:

        excel_file = request.FILES['excel-file']

        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        uploaded_file_url = fs.url(filename)
        wb = openpyxl.load_workbook(excel_file)

        worksheet = wb[wb.sheetnames[0]]

        music_name = [mus['name'] for mus in music_values]
        comp_name = [comp['name'] for comp in comp_values]
        type_name = [t['name'] for t in type_]

       
        i = 0
        for data in worksheet.iter_rows(min_row=1, max_col=5, values_only=True):

            if i == 0:
                i += 1
                continue
            i += 1

<<<<<<< HEAD
            name = data[1]
        
            if name == None:
                break

            elif name not in music_name:
                
                if data[3] not in comp_name or data[4] not in type_name or validate_date(str(data[2])):
                    continue

                t = TypeOfPiece.objects.get(name = data[4])
                c = Compouser.objects.get(name = data[3])

                new_model = {
                    'name': data[1], 
                    'year_written': data[2], 
                    'compousers': c, 
                    'type_of_piece': t
                }

                p = PieceOfMusic(**new_model)
                p.save()

            else:
                for c in music_values:

                    if name != c['name']:
                        continue
                    else:

                        if data[3] not in comp_name or data[4] not in type_name or not validate_date(str(data[2])):
                            continue

                        year_written = data[2]

                        p = PieceOfMusic.objects.get(name = data[1])
                        c = Compouser.objects.get(name = data[3])

                        compouser_id = c.id 

                        t = TypeOfPiece.objects.get(name = data[4])

                        type_of_piece_id = t.id

                        PieceOfMusic.objects.filter(pk=p.id).update(
                            year_written=year_written,
                            compousers_id=c.id, 
                            type_of_piece_id=t.id
                        )

                            
    music_values = PieceOfMusic.objects.order_by(
            'year_written').values('name', 'year_written', 'id')
    index = 0
    
    for mus in music_values:
        m = PieceOfMusic.objects.get(pk=mus['id'])
        music_values[index]['compouser'] = m.compousers.name
        music_values[index]['type_of_piece'] = m.type_of_piece.name
        music_values[index]['year_written'] = str(m.year_written)[:4]
        del music_values[index]['id']
        music_values[index]['id'] = m.id
        index += 1

    music_list = PieceOfMusic.objects.order_by('year_written')
    header = list(music_values[0].keys())[:-1]
    content_music = list(music_values)
    
    return render(request, 'music_search/music.html', {
        'uploaded_file_url': uploaded_file_url,
        'header': header,
        'content_music': content_music,
        'music_list': music_list
    })

class DetailView(LoginRequiredMixin, generic.DetailView):
    model = Period
    template_name = 'music_search/period_detail.html'
    login_url = '/login/'
=======
class CompousersDetailView(generic.DetailView):
    model = Compousers
    template_name = 'music_search/compousers_detail.html'
>>>>>>> 3c85a009d41e0aa88e8966791cf5bf3e6148b68e

    def get_context_data(self, **kwargs):
        context = super(CompousersDetailView, self).get_context_data(**kwargs)
        return context


<<<<<<< HEAD
class CompousersDetailView(LoginRequiredMixin, generic.DetailView):
    model = Compouser
    template_name = 'music_search/compousers_detail.html'
    login_url = '/login/'

    def get_context_data(self, **kwargs):
        context = super(CompousersDetailView, self).get_context_data(**kwargs)
        return context

class MusicDetailView(LoginRequiredMixin, generic.DeleteView):
    model = PieceOfMusic
    template_name = 'music_search/music_detail.html'
    login_url = '/login/'
=======
class PeriodCreate(CreateView):
    model = Period
    form_class = PeriodForm
    template_name = 'music_search/period_create.html'
    #fields = ['time_of_period', 'name', 'descr']
>>>>>>> 3c85a009d41e0aa88e8966791cf5bf3e6148b68e

    def get_context_data(self, **kwargs):
        context = super(MusicDetailView, self).get_context_data(**kwargs)
        music = PieceOfMusic.objects.get(id=self.object.id)
        sheets = music.sheet_set.all()
        sheet = []
        for sh in sheets:
            url = sh.sheet.decode("utf-8")
            if url not in sheet:
                sheet.append(url)
        parts = music.part_set.all()
        audio = []
        for part in parts:
            for audios in part.audios.all():
                url = audios.audio_rec.decode("utf-8")
                if url not in audio:
                    audio.append(url) 
        names = [ part['name'] for part in list(parts.values('name'))]
        context['parts'] = parts
        context['parts_name'] = names
        context['audio'] = audio
        context['sheet'] = sheet
        return context

@method_decorator([login_required, admin_required], name='dispatch')
class PeriodCreate(LoginRequiredMixin, CreateView):
    model = Period
    form_class = PeriodForm
<<<<<<< HEAD
    template_name = 'music_search/period_create.html'
    login_url = '/login/'
    #fields = ['time_of_period', 'name', 'descr']
=======
    template_name = 'music_search/period_edit.html'
>>>>>>> 3c85a009d41e0aa88e8966791cf5bf3e6148b68e

@method_decorator([login_required, admin_required], name='dispatch')
class PeriodUpdate(LoginRequiredMixin, UpdateView):
    model = Period
    form_class = PeriodForm
    template_name = 'music_search/period_edit.html'
    login_url = '/login/'

@method_decorator([login_required, admin_required], name='dispatch')
class PeriodDelete(LoginRequiredMixin, DeleteView):
    model = Period
    template_name = 'music_search/period_delete.html'
    login_url = '/login/'

    def get_success_url(self):
        return reverse('music_search:period')

@method_decorator([login_required, admin_required], name='dispatch')
class CompousersCreate(LoginRequiredMixin, CreateView):
    model = Compouser
    form_class = CompousersForm
    template_name = 'music_search/compousers_create.html'
    login_url = '/login/'

@method_decorator([login_required, admin_required], name='dispatch')
class CompousersUpdate(LoginRequiredMixin, UpdateView):
    model = Compouser
    form_class = CompousersForm
    template_name = 'music_search/compousers_edit.html'
    login_url = '/login/'

@method_decorator([login_required, admin_required], name='dispatch')
class CompousersDelete(LoginRequiredMixin, DeleteView):
    model = Compouser
    template_name = 'music_search/compousers_delete.html'
    login_url = '/login/'

    def get_success_url(self):
        return reverse('music_search:compouser')

@method_decorator([login_required, admin_required], name='dispatch')
class PieceOfMusicCreate(LoginRequiredMixin, CreateView):
    model = PieceOfMusic
    form_class = MusicForm
    template_name = 'music_search/music_create.html'
    login_url = '/login/'

@method_decorator([login_required, admin_required], name='dispatch')
class PieceOfMusicUpdate(LoginRequiredMixin, UpdateView):
    model = PieceOfMusic
    form_class = MusicForm
    template_name = 'music_search/music_edit.html'
    login_url = '/login/'

@method_decorator([login_required, admin_required], name='dispatch')
class PieceOfMusicDelete(LoginRequiredMixin, DeleteView):
    model = PieceOfMusic
    template_name = 'music_search/music_delete.html'
    login_url = '/login/'

    def get_success_url(self):
        return reverse('music_search:music')




class CompousersCreate(CreateView):
    model = Compousers
    form_class = CompousersForm
    template_name = 'music_search/compousers_create.html'
    #fields = ['time_of_period', 'name', 'descr']


class CompousersUpdate(UpdateView):
    model = Compousers
    form_class = CompousersForm
    template_name = 'music_search/compousers_edit.html'


class CompousersDelete(DeleteView):
    model = Compousers
    success_url = reverse_lazy('period-list')
